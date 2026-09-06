"""Keywords + default-replies routes (Strangler Fig P2-B2c).

Mechanically extracted from reply_server.py; behavior-preserving.
Shared models/helpers/state live in app/api/models.py, app/api/common.py and app/api/state.py; reply_server-resident symbols are accessed late-bound (reply_server.X) so runtime rebinds stay visible.
"""

from typing import Any, Dict, List, Optional, Tuple, Callable, Awaitable
from collections import defaultdict
from datetime import datetime, timedelta
import asyncio
import base64
import hashlib
import io
import json
import os
import random
import re
import secrets
import time
import urllib.parse
from urllib.parse import unquote
from urllib import request as urllib_request, error as urllib_error

from PIL import Image, ImageDraw, ImageFilter, ImageFont

from fastapi import (APIRouter, BackgroundTasks, Depends, File, Form, Header,
                     HTTPException, Request, Response, UploadFile, status)
from fastapi.responses import (HTMLResponse, JSONResponse, RedirectResponse,
                               StreamingResponse)
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from loguru import logger
from pydantic import BaseModel

from app.api.models import (
    DefaultReplyIn,
    KeywordIn,
    KeywordWithItemIdIn,
)
import db_manager
import reply_server  # noqa: F401  (late-bound seam: runtime rebinds stay visible)
from utils.image_utils import image_manager
import cookie_manager
import pandas as pd


def create_keywords_router() -> APIRouter:
    router = APIRouter()
    @router.get("/keywords/{cid}")
    def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 直接从数据库获取所有关键词（避免重复计算）
        item_keywords = db_manager.db_manager.get_keywords_with_item_id(cid)

        # 转换为统一格式
        all_keywords = []
        for keyword, reply, item_id in item_keywords:
            all_keywords.append({
                "keyword": keyword,
                "reply": reply,
                "item_id": item_id,
                "type": "item" if item_id else "normal"
            })

        return all_keywords

    @router.get("/keywords-with-item-id/{cid}")
    def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """获取包含商品ID的关键词列表"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 获取包含类型信息的关键词
        keywords = db_manager.db_manager.get_keywords_with_type(cid)

        # 转换为前端需要的格式
        result = []
        for keyword_data in keywords:
            result.append({
                "keyword": keyword_data['keyword'],
                "reply": keyword_data['reply'],
                "item_id": keyword_data['item_id'] or "",
                "type": keyword_data['type'],
                "image_url": keyword_data['image_url'],
                "item_title": keyword_data.get('item_title', '')  # 添加商品名称
            })

        return result

    @router.post("/keywords/{cid}")
    def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            reply_server.log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        kw_list = [(k, v) for k, v in body.keywords.items()]
        reply_server.log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

        cookie_manager.manager.update_keywords(cid, kw_list)
        reply_server.log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
        return {"msg": "updated", "count": len(kw_list)}

    @router.post("/keywords-with-item-id/{cid}")
    def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """更新包含商品ID的关键词列表"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            reply_server.log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证数据格式
        keywords_to_save = []
        keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

        for kw_data in body.keywords:
            keyword = kw_data.get('keyword', '').strip()
            reply = kw_data.get('reply', '').strip()
            item_id = kw_data.get('item_id', '').strip() or None

            if not keyword:
                raise HTTPException(status_code=400, detail="关键词不能为空")

            # 检查当前提交的关键词中是否有重复
            keyword_key = f"{keyword}|{item_id or ''}"
            if keyword_key in keyword_set:
                item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
            keyword_set.add(keyword_key)

            keywords_to_save.append((keyword, reply, item_id))

        # 保存关键词（只保存文本关键词，保留图片关键词）
        try:
            success = db_manager.db_manager.save_text_keywords_only(cid, keywords_to_save)
            if not success:
                raise HTTPException(status_code=500, detail="保存关键词失败")
        except Exception as e:
            error_msg = str(e)

            # 检查是否是图片关键词冲突
            if "已存在（图片关键词）" in error_msg:
                # 直接使用数据库管理器提供的友好错误信息
                raise HTTPException(status_code=400, detail=error_msg)
            elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
                # 尝试从错误信息中提取具体的冲突关键词
                conflict_keyword = None
                conflict_type = None

                # 检查是否是数据库管理器抛出的详细错误
                if "关键词唯一约束冲突" in error_msg:
                    # 解析详细错误信息：关键词唯一约束冲突: Cookie=xxx, 关键词='xxx', 通用关键词/商品ID: xxx
                    import re
                    keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                    if keyword_match:
                        conflict_keyword = keyword_match.group(1)

                    if "通用关键词" in error_msg:
                        conflict_type = "通用关键词"
                    elif "商品ID:" in error_msg:
                        item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                        if item_match:
                            conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

                # 构造用户友好的错误信息
                if conflict_keyword and conflict_type:
                    detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
                elif "keywords.cookie_id, keywords.keyword" in error_msg:
                    detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
                else:
                    detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

                raise HTTPException(status_code=400, detail=detail_msg)
            else:
                reply_server.log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
                raise HTTPException(status_code=500, detail="保存关键词失败")

        reply_server.log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
        return {"msg": "updated", "count": len(keywords_to_save)}

    @router.get("/keywords-export/{cid}")
    def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """导出指定账号的关键词为Excel文件"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        try:
            # 获取关键词数据（包含类型信息）
            keywords = db_manager.db_manager.get_keywords_with_type(cid)

            # 创建DataFrame，只导出文本类型的关键词
            data = []
            for keyword_data in keywords:
                # 只导出文本类型的关键词
                if keyword_data.get('type', 'text') == 'text':
                    data.append({
                        '关键词': keyword_data['keyword'],
                        '商品ID': keyword_data['item_id'] or '',
                        '关键词内容': keyword_data['reply']
                    })

            # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
            if not data:
                df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
            else:
                df = pd.DataFrame(data)

            # 创建Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='关键词数据', index=False)

                # 如果是空模板，添加一些示例说明
                if data == []:
                    worksheet = writer.sheets['关键词数据']
                    # 添加示例数据作为注释（从第2行开始）
                    worksheet['A2'] = '你好'
                    worksheet['B2'] = ''
                    worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                    worksheet['A3'] = '价格'
                    worksheet['B3'] = '123456'
                    worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                    worksheet['A4'] = '发货'
                    worksheet['B4'] = ''
                    worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                    # 设置示例行的样式（浅灰色背景）
                    from openpyxl.styles import PatternFill
                    gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    for row in range(2, 5):
                        for col in range(1, 4):
                            worksheet.cell(row=row, column=col).fill = gray_fill

            output.seek(0)

            # 生成文件名（使用URL编码处理中文）
            from urllib.parse import quote
            if not data:
                filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
            else:
                filename = f"keywords_{cid}_{int(time.time())}.xlsx"
            encoded_filename = quote(filename.encode('utf-8'))

            # 返回文件
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )

        except Exception as e:
            logger.error(f"导出关键词失败: {e}")
            raise HTTPException(status_code=500, detail=f"导出关键词失败: {str(e)}")

    @router.post("/keywords-import/{cid}")
    async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """导入Excel文件中的关键词到指定账号"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 检查文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

        try:
            # 读取Excel文件
            contents = await file.read()
            df = pd.read_excel(io.BytesIO(contents))

            # 检查必要的列
            required_columns = ['关键词', '商品ID', '关键词内容']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

            # 获取现有的文本类型关键词（用于比较更新/新增）
            existing_keywords = db_manager.db_manager.get_keywords_with_type(cid)
            existing_dict = {}
            for keyword_data in existing_keywords:
                # 只考虑文本类型的关键词
                if keyword_data.get('type', 'text') == 'text':
                    keyword = keyword_data['keyword']
                    reply = keyword_data['reply']
                    item_id = keyword_data['item_id']
                    key = f"{keyword}|{item_id or ''}"
                    existing_dict[key] = (keyword, reply, item_id)

            # 处理导入数据
            import_data = []
            update_count = 0
            add_count = 0

            for index, row in df.iterrows():
                keyword = str(row['关键词']).strip()
                item_id = str(row['商品ID']).strip() if pd.notna(row['商品ID']) and str(row['商品ID']).strip() else None
                reply = str(row['关键词内容']).strip()

                if not keyword:
                    continue  # 跳过没有关键词的行

                # 检查是否重复
                key = f"{keyword}|{item_id or ''}"
                if key in existing_dict:
                    # 更新现有关键词
                    update_count += 1
                else:
                    # 新增关键词
                    add_count += 1

                import_data.append((keyword, reply, item_id))

            if not import_data:
                raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

            # 保存到数据库（只影响文本关键词，保留图片关键词）
            success = db_manager.db_manager.save_text_keywords_only(cid, import_data)
            if not success:
                raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

            reply_server.log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

            return {
                "msg": "导入成功",
                "total": len(import_data),
                "added": add_count,
                "updated": update_count
            }

        except pd.errors.EmptyDataError:
            raise HTTPException(status_code=400, detail="Excel文件为空")
        except pd.errors.ParserError:
            raise HTTPException(status_code=400, detail="Excel文件格式错误")
        except Exception as e:
            logger.error(f"导入关键词失败: {e}")
            raise HTTPException(status_code=500, detail=f"导入关键词失败: {str(e)}")

    @router.post("/keywords/{cid}/image")
    async def add_image_keyword(
        cid: str,
        keyword: str = Form(...),
        item_id: str = Form(default=""),
        image: UploadFile = File(...),
        current_user: Dict[str, Any] = Depends(reply_server.get_current_user)
    ):
        """添加图片关键词"""
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查参数
        if not keyword or not keyword.strip():
            raise HTTPException(status_code=400, detail="关键词不能为空")

        if not image or not image.filename:
            raise HTTPException(status_code=400, detail="请选择图片文件")

        # 检查cookie是否属于当前用户
        cookie_details = db_manager.db_manager.get_cookie_details(cid)
        if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
            raise HTTPException(status_code=404, detail="账号不存在或无权限")

        try:
            logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

            # 验证图片文件
            if not image.content_type or not image.content_type.startswith('image/'):
                logger.warning(f"无效的图片文件类型: {image.content_type}")
                raise HTTPException(status_code=400, detail="请上传图片文件")

            # 读取图片数据
            image_data = await image.read()
            logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

            # 保存图片
            image_url = image_manager.save_image(image_data, image.filename)
            if not image_url:
                logger.error("图片保存失败")
                raise HTTPException(status_code=400, detail="图片保存失败")

            logger.info(f"图片保存成功: {image_url}")

            # 先检查关键词是否已存在
            normalized_item_id = item_id if item_id and item_id.strip() else None
            if db_manager.db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
                # 删除已保存的图片
                image_manager.delete_image(image_url)
                if normalized_item_id:
                    raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
                else:
                    raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

            # 保存图片关键词到数据库
            success = db_manager.db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
            if not success:
                # 如果数据库保存失败，删除已保存的图片
                logger.error("数据库保存失败，删除已保存的图片")
                image_manager.delete_image(image_url)
                raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

            reply_server.log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

            return {
                "msg": "图片关键词添加成功",
                "keyword": keyword,
                "image_url": image_url,
                "item_id": item_id or None
            }

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"添加图片关键词失败: {e}")
            raise HTTPException(status_code=500, detail=f"添加图片关键词失败: {str(e)}")

    @router.post("/keywords/{cid}/image-batch")
    async def add_image_keyword_batch(
        cid: str,
        request: Request,
        current_user: Dict[str, Any] = Depends(reply_server.get_current_user)
    ):
        """批量添加图片关键词（使用已上传的图片URL）"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        cookie_details = db_manager.db_manager.get_cookie_details(cid)
        if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
            raise HTTPException(status_code=404, detail="账号不存在或无权限")

        try:
            body = await request.json()
            image_url = body.get('image_url', '').strip()
            keywords = body.get('keywords', [])
            item_ids = body.get('item_ids', [])

            if not image_url:
                raise HTTPException(status_code=400, detail="图片URL不能为空")

            if not keywords or len(keywords) == 0:
                raise HTTPException(status_code=400, detail="关键词列表不能为空")

            # 如果没有商品ID，则使用空字符串（通用关键词）
            if not item_ids or len(item_ids) == 0:
                item_ids = ['']

            logger.info(f"批量添加图片关键词: cid={cid}, keywords={keywords}, item_ids={item_ids}, image_url={image_url}")

            # 检查重复并批量添加
            success_count = 0
            fail_count = 0
            duplicates = []

            for keyword in keywords:
                keyword = keyword.strip()
                if not keyword:
                    continue

                for item_id in item_ids:
                    normalized_item_id = item_id if item_id and item_id.strip() else None

                    # 检查是否重复
                    if db_manager.db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
                        item_id_text = f"（商品ID: {normalized_item_id}）" if normalized_item_id else "（通用关键词）"
                        duplicates.append(f'"{keyword}" {item_id_text}')
                        fail_count += 1
                        continue

                    # 保存图片关键词
                    success = db_manager.db_manager.save_image_keyword(cid, keyword, image_url, normalized_item_id)
                    if success:
                        success_count += 1
                    else:
                        fail_count += 1

            if duplicates:
                reply_server.log_with_user('warning', f"批量添加图片关键词有重复: {cid}, duplicates={duplicates}", current_user)

            reply_server.log_with_user('info', f"批量添加图片关键词完成: {cid}, success={success_count}, fail={fail_count}", current_user)

            return {
                "msg": "批量添加完成",
                "success_count": success_count,
                "fail_count": fail_count,
                "duplicates": duplicates,
                "image_url": image_url
            }

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"批量添加图片关键词失败: {e}")
            raise HTTPException(status_code=500, detail=f"批量添加图片关键词失败: {str(e)}")

    @router.get("/keywords-with-type/{cid}")
    def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """获取包含类型信息的关键词列表"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        cookie_details = db_manager.db_manager.get_cookie_details(cid)
        if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
            raise HTTPException(status_code=404, detail="账号不存在或无权限")

        try:
            keywords = db_manager.db_manager.get_keywords_with_type(cid)
            return keywords
        except Exception as e:
            logger.error(f"获取关键词列表失败: {e}")
            raise HTTPException(status_code=500, detail=f"获取关键词列表失败: {str(e)}")

    @router.delete("/keywords/{cid}/{index}")
    def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """根据索引删除关键词"""
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail="CookieManager 未就绪")

        # 检查cookie是否属于当前用户
        cookie_details = db_manager.db_manager.get_cookie_details(cid)
        if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
            raise HTTPException(status_code=404, detail="账号不存在或无权限")

        try:
            # 先获取要删除的关键词信息（用于删除图片文件）
            keywords = db_manager.db_manager.get_keywords_with_type(cid)
            if 0 <= index < len(keywords):
                keyword_data = keywords[index]

                # 删除关键词
                success = db_manager.db_manager.delete_keyword_by_index(cid, index)
                if not success:
                    raise HTTPException(status_code=400, detail="删除关键词失败")

                # 如果是图片关键词，删除对应的图片文件
                if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                    image_manager.delete_image(keyword_data['image_url'])

                reply_server.log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

                return {"msg": "删除成功"}
            else:
                raise HTTPException(status_code=400, detail="关键词索引无效")

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"删除关键词失败: {e}")
            raise HTTPException(status_code=500, detail=f"删除关键词失败: {str(e)}")

    @router.get("/debug/keywords-table-info")
    def debug_keywords_table_info(admin_user: Dict[str, Any] = Depends(reply_server.require_admin)):
        """调试：检查keywords表结构"""
        try:
            cursor = db_manager.db_manager.conn.cursor()

            # 获取表结构信息
            cursor.execute("PRAGMA table_info(keywords)")
            columns = cursor.fetchall()

            # 获取数据库版本
            cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
            version_result = cursor.fetchone()
            db_version = version_result[0] if version_result else "未知"

            return {
                "db_version": db_version,
                "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
            }
        except Exception as e:
            logger.error(f"检查表结构失败: {e}")
            raise HTTPException(status_code=500, detail=f"检查表结构失败: {str(e)}")

    @router.get('/default-replies/{cid}')
    def get_default_reply(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """获取指定账号的默认回复设置"""
        try:
            # 检查cookie是否属于当前用户
            user_id = current_user['user_id']
            user_cookies = db_manager.db_manager.get_all_cookies(user_id)

            if cid not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限访问该Cookie")

            result = db_manager.db_manager.get_default_reply(cid)
            if result is None:
                # 如果没有设置，返回默认值
                return {'enabled': False, 'reply_content': '', 'reply_once': False}
            return result
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @router.put('/default-replies/{cid}')
    def update_default_reply(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """更新指定账号的默认回复设置"""
        try:
            # 检查cookie是否属于当前用户
            user_id = current_user['user_id']
            user_cookies = db_manager.db_manager.get_all_cookies(user_id)

            if cid not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限操作该Cookie")

            db_manager.db_manager.save_default_reply(cid, reply_data.enabled, reply_data.reply_content, reply_data.reply_once)
            return {'msg': 'default reply updated', 'enabled': reply_data.enabled, 'reply_once': reply_data.reply_once}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/default-replies')
    def get_all_default_replies(current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """获取当前用户所有账号的默认回复设置"""
        try:
            # 只返回当前用户的默认回复设置
            user_id = current_user['user_id']
            user_cookies = db_manager.db_manager.get_all_cookies(user_id)

            all_replies = db_manager.db_manager.get_all_default_replies()
            # 过滤只属于当前用户的回复设置
            user_replies = {cid: reply for cid, reply in all_replies.items() if cid in user_cookies}
            return user_replies
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @router.delete('/default-replies/{cid}')
    def delete_default_reply(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """删除指定账号的默认回复设置"""
        try:
            # 检查cookie是否属于当前用户
            user_id = current_user['user_id']
            user_cookies = db_manager.db_manager.get_all_cookies(user_id)

            if cid not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限操作该Cookie")

            success = db_manager.db_manager.delete_default_reply(cid)
            if success:
                return {'msg': 'default reply deleted'}
            else:
                raise HTTPException(status_code=400, detail='删除失败')
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @router.post('/default-replies/{cid}/clear-records')
    def clear_default_reply_records(cid: str, current_user: Dict[str, Any] = Depends(reply_server.get_current_user)):
        """清空指定账号的默认回复记录"""
        try:
            # 检查cookie是否属于当前用户
            user_id = current_user['user_id']
            user_cookies = db_manager.db_manager.get_all_cookies(user_id)

            if cid not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限操作该Cookie")

            db_manager.db_manager.clear_default_reply_records(cid)
            return {'msg': 'default reply records cleared'}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    return router
