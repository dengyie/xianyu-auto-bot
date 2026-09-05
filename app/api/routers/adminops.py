"""Admin / ops / logs / backup / update / files / groups / blacklist routes (P2-B5).

Mechanically extracted from reply_server.py; behavior-preserving.
External (reply_server) symbols resolve via ctx at request time - see app/api/state.py.
"""

from typing import Any, Dict, List, Optional, Tuple, Callable, Awaitable
from collections import defaultdict
from datetime import datetime, timedelta
import asyncio
import base64
import hashlib
import io
import json
import os
import random
import re
import secrets
import time
import urllib.parse
from urllib.parse import unquote
from urllib import request as urllib_request, error as urllib_error

from PIL import Image, ImageDraw, ImageFilter, ImageFont

from fastapi import (APIRouter, BackgroundTasks, Depends, File, Form, Header,
                     HTTPException, Request, Response, UploadFile, status)
from fastapi.responses import (HTMLResponse, JSONResponse, RedirectResponse,
                               StreamingResponse)
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from loguru import logger
from pydantic import BaseModel


def create_admin_ops_router(ctx) -> APIRouter:
    router = APIRouter()
    @router.get("/ai-reply-settings/{cookie_id}")
    def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """获取指定账号的AI回复设置"""
        try:
            # 检查cookie是否属于当前用户
            user_id = current_user['user_id']
            from db_manager import db_manager
            user_cookies = ctx.db_manager.get_all_cookies(user_id)

            if cookie_id not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限访问该Cookie")

            settings = ctx.db_manager.get_ai_reply_settings(cookie_id)
            return settings
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"获取AI回复设置异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.put("/ai-reply-settings/{cookie_id}")
    def update_ai_reply_settings(cookie_id: str, settings: ctx.AIReplySettings, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """更新指定账号的AI回复设置"""
        try:
            # 检查cookie是否属于当前用户
            user_id = current_user['user_id']
            from db_manager import db_manager
            user_cookies = ctx.db_manager.get_all_cookies(user_id)

            if cookie_id not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限操作该Cookie")

            # 检查账号是否存在
            if ctx.cookie_manager.manager is None:
                raise HTTPException(status_code=500, detail='CookieManager 未就绪')

            # 保存设置
            settings_dict = settings.model_dump()
            success = ctx.db_manager.save_ai_reply_settings(cookie_id, settings_dict)

            if success:

                # 如果启用了AI回复，记录日志
                if settings.ai_enabled:
                    logger.info(f"账号 {cookie_id} 启用AI回复")
                else:
                    logger.info(f"账号 {cookie_id} 禁用AI回复")

                return {"message": "AI回复设置更新成功"}
            else:
                raise HTTPException(status_code=400, detail="更新失败")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"更新AI回复设置异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.get("/ai-reply-settings")
    def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """获取当前用户所有账号的AI回复设置"""
        try:
            # 只返回当前用户的AI回复设置
            user_id = current_user['user_id']
            from db_manager import db_manager
            user_cookies = ctx.db_manager.get_all_cookies(user_id)

            all_settings = ctx.db_manager.get_all_ai_reply_settings()
            # 过滤只属于当前用户的设置
            user_settings = {cid: settings for cid, settings in all_settings.items() if cid in user_cookies}
            return user_settings
        except Exception as e:
            logger.error(f"获取所有AI回复设置异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.get("/ai-config-presets")
    def list_ai_config_presets(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """获取当前用户的AI配置预设列表"""
        try:
            user_id = current_user['user_id']
            from db_manager import db_manager
            presets = ctx.db_manager.get_ai_config_presets(user_id)
            return presets
        except Exception as e:
            logger.error(f"获取AI配置预设列表异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.post("/ai-config-presets")
    def save_ai_config_preset(
        preset: ctx.AIConfigPreset,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user)
    ):
        """创建或更新AI配置预设"""
        try:
            user_id = current_user['user_id']
            from db_manager import db_manager

            # 检查预设数量上限
            existing = ctx.db_manager.get_ai_config_presets(user_id)
            existing_names = [p['preset_name'] for p in existing]
            if preset.preset_name not in existing_names and len(existing) >= 20:
                raise HTTPException(status_code=400, detail="预设数量已达上限（最多20个）")

            preset_id = ctx.db_manager.save_ai_config_preset(
                user_id=user_id,
                preset_name=preset.preset_name,
                model_name=preset.model_name,
                api_key=preset.api_key,
                base_url=preset.base_url,
                api_type=preset.api_type
            )
            return {"message": "预设保存成功", "preset_id": preset_id}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"保存AI配置预设异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.delete("/ai-config-presets/{preset_id}")
    def delete_ai_config_preset(
        preset_id: int,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user)
    ):
        """删除AI配置预设"""
        try:
            user_id = current_user['user_id']
            from db_manager import db_manager
            deleted = ctx.db_manager.delete_ai_config_preset(user_id, preset_id)
            if not deleted:
                raise HTTPException(status_code=404, detail="预设不存在或无权删除")
            return {"message": "预设删除成功"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"删除AI配置预设异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.post("/ai-reply-test/{cookie_id}")
    def test_ai_reply(cookie_id: str, test_data: dict, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """测试AI回复功能"""
        try:
            user_id = current_user['user_id']
            from db_manager import db_manager
            user_cookies = ctx.db_manager.get_all_cookies(user_id)

            if cookie_id not in user_cookies:
                raise HTTPException(status_code=403, detail="无权限操作该Cookie")

            # 检查账号是否存在
            if ctx.cookie_manager.manager is None:
                raise HTTPException(status_code=500, detail='CookieManager 未就绪')

            if cookie_id not in ctx.cookie_manager.manager.cookies:
                raise HTTPException(status_code=404, detail='账号不存在')

            # 检查是否启用AI回复
            if not ctx.ai_reply_engine.is_ai_enabled(cookie_id):
                raise HTTPException(status_code=400, detail='该账号未启用AI回复')

            # 构造测试数据
            test_message = test_data.get('message', '你好')
            test_item_info = {
                'title': test_data.get('item_title', '测试商品'),
                'price': test_data.get('item_price', 100),
                'desc': test_data.get('item_desc', '这是一个测试商品')
            }

            # 生成测试回复（跳过去抖等待）
            reply = ctx.ai_reply_engine.generate_reply(
                message=test_message,
                item_info=test_item_info,
                chat_id=f"test_{int(time.time())}",
                cookie_id=cookie_id,
                user_id="test_user",
                item_id="test_item",
                skip_wait=True
            )

            if reply:
                return {"message": "测试成功", "reply": reply}
            else:
                raise HTTPException(status_code=400, detail="AI回复生成失败")

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"测试AI回复异常: {e}")
            raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")

    @router.get('/api/task-logs')
    def get_task_logs(
        task_type: str = 'all',
        cookie_id: str = None,
        limit: int = 100,
        offset: int = 0,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        """查询系统日志页的统一任务日志。"""
        try:
            safe_limit = ctx._normalize_task_log_limit(limit)
            safe_offset = ctx._normalize_task_log_offset(offset)
            requested_type = str(task_type or 'all').strip() or 'all'
            if requested_type not in {'all', *ctx.TASK_LOG_TYPE_LABELS.keys()}:
                requested_type = 'all'

            scoped_cookie_id = None
            if cookie_id:
                scoped_cookie_id = ctx._ensure_cookie_access(cookie_id, current_user)

            logs: List[Dict[str, Any]] = []

            if requested_type in {'all', 'auto_comment'}:
                logs.extend(
                    ctx._normalize_task_log_row(log, 'auto_comment')
                    for log in ctx.db_manager.get_scheduled_rate_logs(
                        user_id=current_user['user_id'],
                        cookie_id=scoped_cookie_id,
                        limit=safe_limit,
                        offset=0,
                    )
                )

            # 本地未接入求小红花，跳过 auto_red_flower 源

            generic_types = {'item_polish', 'login_renew', 'cookie_refresh', 'other_task'}
            if requested_type == 'all':
                generic_task_type = None
            elif requested_type in generic_types:
                generic_task_type = requested_type
            else:
                generic_task_type = '__skip__'

            if generic_task_type != '__skip__':
                generic_logs = ctx.db_manager.get_scheduled_task_logs(
                    user_id=current_user['user_id'],
                    cookie_id=scoped_cookie_id,
                    task_type=generic_task_type,
                    limit=safe_limit,
                    offset=0,
                )
                logs.extend(
                    ctx._normalize_task_log_row(log, log.get('task_type') or 'other_task')
                    for log in generic_logs
                )

            if requested_type in {'all', 'login_renew', 'cookie_refresh', 'other_task'}:
                logs.extend(ctx._load_risk_task_logs(current_user, requested_type, scoped_cookie_id, safe_limit))

            logs.sort(key=ctx._task_log_created_at_sort_value, reverse=True)
            page = logs[safe_offset:safe_offset + safe_limit]
            return {"success": True, "data": page, "total": len(logs)}
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"查询统一任务日志失败: {str(e)}", current_user)
            raise HTTPException(status_code=500, detail=f"查询统一任务日志失败: {str(e)}")

    @router.get('/api/auto-comment/logs')
    def get_auto_comment_logs(
        cookie_id: str = None,
        limit: int = 100,
        offset: int = 0,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        """查询自动评价执行日志。"""
        try:
            if cookie_id:
                cookie_id = ctx._ensure_cookie_access(cookie_id, current_user)
            logs = ctx.db_manager.get_scheduled_rate_logs(
                user_id=current_user['user_id'],
                cookie_id=cookie_id,
                limit=limit,
                offset=offset,
            )
            return {"success": True, "data": logs}
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"查询自动评价日志失败: {str(e)}", current_user)
            raise HTTPException(status_code=500, detail=f"查询自动评价日志失败: {str(e)}")

    @router.post('/api/auto-comment/batch-rate')
    async def batch_rate_historical_orders(
        request: ctx.AutoCommentBatchRateRequest,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        """从闲鱼待评价列表拉取历史订单并批量补评价。"""
        try:
            from utils.rate_service import RateService, fetch_merchant_rate_list

            raw_ids = request.cookie_ids if request.cookie_ids is not None else request.account_ids
            account_ids = list(dict.fromkeys(
                str(account_id or '').strip()
                for account_id in (raw_ids or [])
                if str(account_id or '').strip()
            ))
            if not account_ids:
                raise HTTPException(status_code=400, detail='请选择账号')

            page_size = max(1, min(int(request.page_size or 100), 100))
            batch_id = f"manual_history_rate_{ctx.uuid.uuid4()}"
            details = []
            stats = {
                'batch_id': batch_id,
                'total_accounts': len(account_ids),
                'success_accounts': 0,
                'total_pending': 0,
                'total_rated': 0,
                'total_failed': 0,
                'total_skipped': 0,
            }

            for raw_cookie_id in account_ids:
                account_result = {
                    'account_id': raw_cookie_id,
                    'success': False,
                    'rated_count': 0,
                    'failed_count': 0,
                    'skipped_count': 0,
                    'total_pending': 0,
                    'message': '',
                }
                try:
                    cookie_id = ctx._ensure_cookie_access(raw_cookie_id, current_user)
                    account_result['account_id'] = cookie_id

                    if not ctx.db_manager.get_auto_comment(cookie_id):
                        account_result['message'] = '未开启自动好评'
                        account_result['skipped_count'] += 1
                        stats['total_skipped'] += 1
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id, cookie_id, status='skipped', message='历史补评价跳过：未开启自动好评'
                        )
                        details.append(account_result)
                        continue

                    template = ctx.db_manager.get_active_comment_template(cookie_id)
                    feedback = str((template or {}).get('content') or '').strip()
                    if not feedback:
                        account_result['message'] = '未设置激活的好评模板'
                        account_result['skipped_count'] += 1
                        stats['total_skipped'] += 1
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id, cookie_id, status='missing_template', message='历史补评价跳过：未设置激活的好评模板'
                        )
                        details.append(account_result)
                        continue

                    cookie_string = ctx.db_manager.get_cookie(cookie_id)
                    if not cookie_string:
                        account_result['message'] = '账号 Cookie 为空或不存在'
                        account_result['failed_count'] += 1
                        stats['total_failed'] += 1
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id, cookie_id, status='cookie_expired', message='历史补评价失败：账号 Cookie 为空或不存在'
                        )
                        details.append(account_result)
                        continue

                    list_result = await fetch_merchant_rate_list(
                        cookie_string=cookie_string,
                        account_id=cookie_id,
                        page=1,
                        page_size=page_size,
                        max_retries=3,
                    )
                    if not list_result.get('success'):
                        status = 'cookie_expired' if list_result.get('session_expired') else 'failed'
                        message = f"获取待评价列表失败: {list_result.get('message') or '未知错误'}"
                        account_result['message'] = message
                        account_result['failed_count'] += 1
                        stats['total_failed'] += 1
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id=batch_id,
                            cookie_id=cookie_id,
                            status=status,
                            message=message,
                            raw_response=list_result.get('raw') or list_result,
                        )
                        details.append(account_result)
                        continue

                    pending_items = list_result.get('items') or []
                    if not isinstance(pending_items, list):
                        pending_items = []
                    account_result['total_pending'] = len(pending_items)
                    stats['total_pending'] += len(pending_items)

                    if not pending_items:
                        account_result['success'] = True
                        account_result['message'] = '没有待评价订单'
                        stats['success_accounts'] += 1
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id, cookie_id, status='skipped', message='历史补评价：没有待评价订单'
                        )
                        details.append(account_result)
                        continue

                    current_cookie = str(list_result.get('cookies_str') or cookie_string)
                    for item in pending_items:
                        meta = ctx._extract_merchant_rate_item_meta(item if isinstance(item, dict) else {})
                        order_id = ctx._extract_merchant_rate_order_id(item if isinstance(item, dict) else {})
                        if not order_id:
                            account_result['failed_count'] += 1
                            stats['total_failed'] += 1
                            ctx.db_manager.add_scheduled_rate_log(
                                batch_id=batch_id,
                                cookie_id=cookie_id,
                                item_id=meta.get('item_id') or None,
                                buyer_id=meta.get('buyer_id') or None,
                                buyer_nick=meta.get('buyer_nick') or None,
                                comment=feedback,
                                status='failed',
                                message='待评价列表项缺少订单号',
                                raw_response=item,
                            )
                            continue

                        rate_service = RateService(current_cookie, account_id=cookie_id)
                        rate_result = await rate_service.rate_buyer(order_id, feedback=feedback)
                        if rate_service.cookie_string and rate_service.cookie_string != current_cookie:
                            current_cookie = rate_service.cookie_string

                        status = 'already_rated' if rate_result.get('already_rated') else (
                            'success' if rate_result.get('success') else (
                                'cookie_expired' if rate_result.get('session_expired') else 'failed'
                            )
                        )
                        message = str(rate_result.get('message') or '')
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id=batch_id,
                            cookie_id=cookie_id,
                            order_id=order_id,
                            item_id=meta.get('item_id') or None,
                            buyer_id=meta.get('buyer_id') or None,
                            buyer_nick=meta.get('buyer_nick') or None,
                            comment=feedback,
                            status=status,
                            message=message,
                            raw_response=rate_result.get('raw') or rate_result,
                        )

                        if rate_result.get('success'):
                            account_result['rated_count'] += 1
                            stats['total_rated'] += 1
                            ctx.db_manager.mark_order_rated(order_id, True)
                        else:
                            account_result['failed_count'] += 1
                            stats['total_failed'] += 1
                            ctx.db_manager.mark_order_rated(order_id, False, message)

                        await asyncio.sleep(1)

                    account_result['success'] = True
                    account_result['message'] = (
                        f"评价完成: 成功 {account_result['rated_count']} 笔，"
                        f"失败 {account_result['failed_count']} 笔"
                    )
                    stats['success_accounts'] += 1
                    details.append(account_result)
                except HTTPException as exc:
                    account_result['message'] = str(exc.detail or '账号无权限或不存在')
                    account_result['failed_count'] += 1
                    stats['total_failed'] += 1
                    details.append(account_result)
                except Exception as exc:
                    logger.error(f"[历史补评价] 账号 {raw_cookie_id} 处理异常: {exc}")
                    account_result['message'] = f"处理异常: {str(exc)}"
                    account_result['failed_count'] += 1
                    stats['total_failed'] += 1
                    try:
                        ctx.db_manager.add_scheduled_rate_log(
                            batch_id, raw_cookie_id, status='failed', message=account_result['message']
                        )
                    except Exception:
                        pass
                    details.append(account_result)

            message = (
                f"历史补评价完成: {stats['success_accounts']}/{stats['total_accounts']} 个账号处理成功，"
                f"共评价 {stats['total_rated']} 笔，失败 {stats['total_failed']} 笔"
            )
            ctx.log_with_user('info', message, current_user)
            return {
                'success': True,
                'message': message,
                'data': {
                    **stats,
                    'details': details,
                },
            }
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"历史补评价失败: {str(e)}", current_user)
            raise HTTPException(status_code=500, detail=f"历史补评价失败: {str(e)}")

    @router.get("/logs")
    async def get_logs(lines: int = 200, level: str = None, source: str = None, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """获取实时系统日志"""
        try:
            # 获取文件日志收集器
            collector = ctx.get_file_log_collector()

            # 获取日志
            logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

            return {"success": True, "logs": logs}

        except Exception as e:
            return {"success": False, "message": f"获取日志失败: {str(e)}", "logs": []}

    @router.get("/risk-control-logs")
    async def get_risk_control_logs(
        cookie_id: str = None,
        processing_status: str = None,
        event_type: str = None,
        trigger_scene: str = None,
        session_id: str = None,
        result_code: str = None,
        date_from: str = None,
        date_to: str = None,
        limit: int = 100,
        offset: int = 0,
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        """获取风控日志（管理员专用）"""
        try:
            ctx.log_with_user(
                'info',
                f"查询风控日志: cookie_id={cookie_id}, processing_status={processing_status}, event_type={event_type}, trigger_scene={trigger_scene}, session_id={session_id}, result_code={result_code}, date_from={date_from}, date_to={date_to}, limit={limit}, offset={offset}",
                admin_user,
            )

            # 获取风控日志
            logs = ctx.db_manager.get_risk_control_logs(
                cookie_id=cookie_id,
                processing_status=processing_status,
                event_type=event_type,
                trigger_scene=trigger_scene,
                session_id=session_id,
                result_code=result_code,
                date_from=date_from,
                date_to=date_to,
                limit=limit,
                offset=offset
            )
            total_count = ctx.db_manager.get_risk_control_logs_count(
                cookie_id=cookie_id,
                processing_status=processing_status,
                event_type=event_type,
                trigger_scene=trigger_scene,
                session_id=session_id,
                result_code=result_code,
                date_from=date_from,
                date_to=date_to,
            )

            ctx.log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

            return {
                "success": True,
                "data": logs,
                "total": total_count,
                "limit": limit,
                "offset": offset
            }

        except Exception as e:
            ctx.log_with_user('error', f"获取风控日志失败: {str(e)}", admin_user)
            return {
                "success": False,
                "message": f"获取风控日志失败: {str(e)}",
                "data": [],
                "total": 0
            }

    @router.get("/logs/stats")
    async def get_log_stats(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """获取日志统计信息"""
        try:
            collector = ctx.get_file_log_collector()
            stats = collector.get_stats()

            return {"success": True, "stats": stats}

        except Exception as e:
            return {"success": False, "message": f"获取日志统计失败: {str(e)}", "stats": {}}

    @router.post("/logs/clear")
    async def clear_logs(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """清空日志"""
        try:
            collector = ctx.get_file_log_collector()
            collector.clear_logs()

            return {"success": True, "message": "日志已清空"}

        except Exception as e:
            return {"success": False, "message": f"清空日志失败: {str(e)}"}

    @router.get("/admin/slider-verification-stats")
    async def get_slider_verification_stats(
        cookie_id: str = None,
        range_key: str = 'all',
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        """获取当前系统用户下的滑块验证统计。"""
        try:
            user_id = admin_user['user_id']
            user_cookie_ids = sorted(ctx.db_manager.get_all_cookies(user_id).keys())
            normalized_range = str(range_key or '').strip().lower()
            if normalized_range not in {'today', '7d', 'all'}:
                normalized_range = 'all'
            range_label = {
                'today': '当日',
                '7d': '近 7 天',
                'all': '所有',
            }[normalized_range]

            if cookie_id:
                if cookie_id not in user_cookie_ids:
                    return {
                        'success': True,
                        'data': {
                            **ctx._empty_slider_session_stats(),
                            'scope_label': cookie_id,
                            'selected_cookie_id': cookie_id,
                            'selected_range': normalized_range,
                            'range_label': range_label,
                            'summary_text': '暂无滑块验证记录' if normalized_range == 'all' else f'{range_label}暂无滑块验证记录',
                        }
                    }
                target_cookie_ids = [cookie_id]
                scope_label = cookie_id
            else:
                target_cookie_ids = user_cookie_ids
                scope_label = '全部账号'

            stats = ctx.db_manager.get_slider_verification_session_stats(target_cookie_ids, range_key=normalized_range)
            stats.update({
                'scope_label': scope_label,
                'selected_cookie_id': cookie_id or '',
            })

            ctx.log_with_user(
                'info',
                f"获取滑块验证统计成功: scope={scope_label}, range={range_label}, sessions={stats['total_sessions']}, success={stats['success_count']}, failure={stats['failure_count']}",
                admin_user,
            )

            return {
                'success': True,
                'data': stats,
            }
        except Exception as e:
            ctx.log_with_user('error', f"获取滑块验证统计失败: {str(e)}", admin_user)
            return {
                'success': False,
                'message': f'获取滑块验证统计失败: {str(e)}',
                'data': ctx._empty_slider_session_stats(),
            }

    @router.delete("/admin/risk-control-logs/{log_id}")
    async def delete_risk_control_log(
        log_id: int,
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        """删除风控日志记录（管理员专用）"""
        try:
            ctx.log_with_user('info', f"删除风控日志记录: {log_id}", admin_user)

            success = ctx.db_manager.delete_risk_control_log(log_id)

            if success:
                ctx.log_with_user('info', f"风控日志删除成功: {log_id}", admin_user)
                return {"success": True, "message": "删除成功"}
            else:
                ctx.log_with_user('warning', f"风控日志删除失败: {log_id}", admin_user)
                return {"success": False, "message": "删除失败，记录可能不存在"}

        except Exception as e:
            ctx.log_with_user('error', f"删除风控日志失败: {log_id} - {str(e)}", admin_user)
            return {"success": False, "message": f"删除失败: {str(e)}"}

    @router.get('/admin/users')
    def get_all_users(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """获取所有用户信息（管理员专用）"""
        from db_manager import db_manager
        try:
            ctx.log_with_user('info', "查询所有用户信息", admin_user)
            users = ctx.db_manager.get_all_users()

            # 为每个用户添加统计信息
            for user in users:
                user_id = user['id']
                # 统计用户的Cookie数量
                user_cookies = ctx.db_manager.get_all_cookies(user_id)
                user['cookie_count'] = len(user_cookies)

                # 统计用户的卡券数量
                user_cards = ctx.db_manager.get_all_cards(user_id)
                user['card_count'] = len(user_cards) if user_cards else 0

                # 隐藏密码字段
                if 'password_hash' in user:
                    del user['password_hash']  
    # ??????????????????????????

            ctx.log_with_user('info', f"返回用户信息，共 {len(users)} 个用户", admin_user)
            return {"users": users}
        except Exception as e:
            ctx.log_with_user('error', f"获取用户信息失败: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.delete('/admin/users/{user_id}')
    def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """删除用户（管理员专用）"""
        from db_manager import db_manager
        try:
            # 不能删除管理员自己
            if user_id == admin_user['user_id']:
                ctx.log_with_user('warning', "尝试删除管理员自己", admin_user)
                raise HTTPException(status_code=400, detail="不能删除管理员自己")

            # 获取要删除的用户信息
            user_to_delete = ctx.db_manager.get_user_by_id(user_id)
            if not user_to_delete:
                raise HTTPException(status_code=404, detail="用户不存在")

            ctx.log_with_user('info', f"准备删除用户: {user_to_delete['username']} (ID: {user_id})", admin_user)

            # 删除用户及其相关数据
            success = ctx.db_manager.delete_user_and_data(user_id)

            if success:
                ctx.audit_event(
                    category="admin",
                    action="admin_user_delete",
                    status="success",
                    actor=admin_user,
                    resource_type="user",
                    resource_id=user_id,
                    message="Admin deleted user",
                    details={
                        "target_username": user_to_delete.get("username"),
                        "target_user_id": user_id,
                    },
                )
                ctx.log_with_user('info', f"用户删除成功: {user_to_delete['username']} (ID: {user_id})", admin_user)
                return {"message": f"用户 {user_to_delete['username']} 删除成功"}
            else:
                ctx.log_with_user('error', f"用户删除失败: {user_to_delete['username']} (ID: {user_id})", admin_user)
                raise HTTPException(status_code=400, detail="删除失败")
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"删除用户异常: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.put('/admin/users/{user_id}/admin-status')
    def update_user_admin_status(user_id: int, is_admin: bool, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """更新用户管理员状态（管理员专用）"""
        from db_manager import db_manager
        try:
            # 获取目标用户信息
            target_user = ctx.db_manager.get_user_by_id(user_id)
            if not target_user:
                raise HTTPException(status_code=404, detail="用户不存在")

            # 不能修改自己的管理员状态（防止误操作导致没有管理员）
            if user_id == admin_user['user_id']:
                ctx.log_with_user('warning', "尝试修改自己的管理员状态", admin_user)
                raise HTTPException(status_code=400, detail="不能修改自己的管理员状态")

            ctx.log_with_user('info', f"准备{'设置' if is_admin else '取消'}{target_user['username']}的管理员权限", admin_user)

            # 更新管理员状态
            success = ctx.db_manager.update_user_admin_status(user_id, is_admin)

            if success:
                action = "设置为管理员" if is_admin else "取消管理员权限"
                removed_tokens = ctx._remove_session_tokens_for_user(user_id)
                ctx.audit_event(
                    category="admin",
                    action="admin_user_status_update",
                    status="success",
                    actor=admin_user,
                    resource_type="user",
                    resource_id=user_id,
                    message="Admin updated user admin status",
                    details={
                        "target_username": target_user.get("username"),
                        "is_admin": is_admin,
                        "revoked_sessions": removed_tokens,
                    },
                )
                ctx.log_with_user('info', f"用户 {target_user['username']} 已{action}", admin_user)
                return {
                    "success": True,
                    "message": f"用户 {target_user['username']} 已{action}",
                    "user_id": user_id,
                    "is_admin": is_admin,
                    "revoked_sessions": removed_tokens
                }
            else:
                ctx.log_with_user('error', f"更新用户管理员状态失败: {target_user['username']}", admin_user)
                raise HTTPException(status_code=400, detail="更新失败")
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"更新用户管理员状态异常: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/admin/risk-control-logs')
    async def get_admin_risk_control_logs(
        cookie_id: str = None,
        processing_status: str = None,
        event_type: str = None,
        trigger_scene: str = None,
        session_id: str = None,
        result_code: str = None,
        date_from: str = None,
        date_to: str = None,
        limit: int = 100,
        offset: int = 0,
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        """获取风控日志（管理员专用）"""
        try:
            ctx.log_with_user(
                'info',
                f"查询风控日志: cookie_id={cookie_id}, processing_status={processing_status}, event_type={event_type}, trigger_scene={trigger_scene}, session_id={session_id}, result_code={result_code}, date_from={date_from}, date_to={date_to}, limit={limit}, offset={offset}",
                admin_user,
            )

            # 获取风控日志
            logs = ctx.db_manager.get_risk_control_logs(
                cookie_id=cookie_id,
                processing_status=processing_status,
                event_type=event_type,
                trigger_scene=trigger_scene,
                session_id=session_id,
                result_code=result_code,
                date_from=date_from,
                date_to=date_to,
                limit=limit,
                offset=offset
            )
            total_count = ctx.db_manager.get_risk_control_logs_count(
                cookie_id=cookie_id,
                processing_status=processing_status,
                event_type=event_type,
                trigger_scene=trigger_scene,
                session_id=session_id,
                result_code=result_code,
                date_from=date_from,
                date_to=date_to,
            )

            ctx.log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

            return {
                "success": True,
                "data": logs,
                "total": total_count,
                "limit": limit,
                "offset": offset
            }

        except Exception as e:
            ctx.log_with_user('error', f"查询风控日志失败: {str(e)}", admin_user)
            return {"success": False, "message": f"查询失败: {str(e)}", "data": [], "total": 0}

    @router.get('/admin/cookies')
    def get_admin_cookies(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """获取所有Cookie信息（管理员专用）"""
        try:
            ctx.log_with_user('info', "查询所有Cookie信息", admin_user)

            if ctx.cookie_manager.manager is None:
                return {
                    "success": True,
                    "cookies": [],
                    "message": "CookieManager 未就绪"
                }

            # 获取所有用户的cookies
            from db_manager import db_manager
            all_users = ctx.db_manager.get_all_users()
            all_cookies = []

            for user in all_users:
                user_id = user['id']
                user_cookies = ctx.db_manager.get_all_cookies(user_id)
                for cookie_id, cookie_value in user_cookies.items():
                    # 获取cookie详细信息
                    cookie_details = ctx.db_manager.get_cookie_details(cookie_id)
                    cookie_info = {
                        'cookie_id': cookie_id,
                        'user_id': user_id,
                        'username': user['username'],
                        'nickname': cookie_details.get('remark', '') if cookie_details else '',
                        'enabled': ctx.cookie_manager.manager.get_cookie_status(cookie_id)
                    }
                    all_cookies.append(cookie_info)

            ctx.log_with_user('info', f"获取到 {len(all_cookies)} 个Cookie", admin_user)
            return {
                "success": True,
                "cookies": all_cookies,
                "total": len(all_cookies)
            }

        except Exception as e:
            ctx.log_with_user('error', f"获取Cookie信息失败: {str(e)}", admin_user)
            return {
                "success": False,
                "cookies": [],
                "message": f"获取失败: {str(e)}"
            }

    @router.get('/admin/audit-logs')
    def get_audit_logs(
        limit: int = 100,
        category: str = None,
        action: str = None,
        status: str = None,
        actor_user_id: int = None,
        resource_type: str = None,
        resource_id: str = None,
        admin_user: Dict[str, Any] = Depends(ctx.require_admin),
    ):
        """Return structured audit logs for administrators."""
        try:
            logs = ctx.db_manager.get_audit_logs(
                limit=limit,
                category=category,
                action=action,
                status=status,
                actor_user_id=actor_user_id,
                resource_type=resource_type,
                resource_id=resource_id,
            )
            ctx.audit_event(
                category="admin",
                action="audit_log_query",
                status="success",
                actor=admin_user,
                resource_type="audit_logs",
                message="Admin queried audit logs",
                details={
                    "limit": limit,
                    "category": category,
                    "action": action,
                    "status": status,
                    "actor_user_id": actor_user_id,
                    "resource_type": resource_type,
                    "resource_id": resource_id,
                    "result_count": len(logs),
                },
            )
            return {"success": True, "logs": logs, "total": len(logs)}
        except Exception as e:
            ctx.audit_event(
                category="admin",
                action="audit_log_query",
                status="error",
                actor=admin_user,
                resource_type="audit_logs",
                message="Admin audit log query failed",
                details={"error": str(e)},
            )
            raise HTTPException(status_code=500, detail="审计日志查询失败")

    @router.get('/admin/logs')
    def get_system_logs(admin_user: Dict[str, Any] = Depends(ctx.require_admin),
                       lines: int = 100,
                       level: str = None):
        """获取系统日志（管理员专用）"""
        import os
        import glob
        from datetime import datetime

        try:
            ctx.log_with_user('info', f"查询系统日志，行数: {lines}, 级别: {level}", admin_user)

            # 查找日志文件
            log_files = glob.glob("logs/xianyu_*.log")
            logger.info(f"找到日志文件: {log_files}")

            if not log_files:
                logger.warning("未找到日志文件")
                return {"logs": [], "message": "未找到日志文件", "success": False}

            # 获取最新的日志文件
            latest_log_file = max(log_files, key=os.path.getctime)
            logger.info(f"使用最新日志文件: {latest_log_file}")

            logs = []
            try:
                with open(latest_log_file, 'r', encoding='utf-8') as f:
                    all_lines = f.readlines()
                    logger.info(f"读取到 {len(all_lines)} 行日志")

                    # 如果指定了日志级别，进行过滤
                    if level:
                        filtered_lines = [line for line in all_lines if f"| {level.upper()} |" in line]
                        logger.info(f"按级别 {level} 过滤后剩余 {len(filtered_lines)} 行")
                    else:
                        filtered_lines = all_lines

                    # 获取最后N行
                    recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                    logger.info(f"取最后 {len(recent_lines)} 行日志")

                    for line in recent_lines:
                        logs.append(line.strip())

            except Exception as e:
                logger.error(f"读取日志文件失败: {str(e)}")
                ctx.log_with_user('error', f"读取日志文件失败: {str(e)}", admin_user)
                return {"logs": [], "message": f"读取日志文件失败: {str(e)}", "success": False}

            ctx.log_with_user('info', f"返回日志记录 {len(logs)} 条", admin_user)
            logger.info(f"成功返回 {len(logs)} 条日志记录")

            return {
                "logs": logs,
                "log_file": latest_log_file,
                "total_lines": len(logs),
                "success": True
            }

        except Exception as e:
            logger.error(f"获取系统日志失败: {str(e)}")
            ctx.log_with_user('error', f"获取系统日志失败: {str(e)}", admin_user)
            return {"logs": [], "message": f"获取系统日志失败: {str(e)}", "success": False}

    @router.get('/admin/log-files')
    def list_log_files(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """列出所有可用的系统日志文件"""
        import os
        import glob
        from datetime import datetime

        try:
            ctx.log_with_user('info', "查询日志文件列表", admin_user)

            log_dir = "logs"
            if not os.path.exists(log_dir):
                logger.warning("日志目录不存在")
                return {"success": True, "files": []}

            log_pattern = os.path.join(log_dir, "xianyu_*.log")
            log_files = glob.glob(log_pattern)

            files_info = []
            for file_path in log_files:
                try:
                    stat_info = os.stat(file_path)
                    files_info.append({
                        "name": os.path.basename(file_path),
                        "size": stat_info.st_size,
                        "modified_at": datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                        "modified_ts": stat_info.st_mtime
                    })
                except OSError as e:
                    logger.warning(f"读取日志文件信息失败 {file_path}: {e}")

            # 按修改时间倒序排序
            files_info.sort(key=lambda item: item.get("modified_ts", 0), reverse=True)

            logger.info(f"返回日志文件列表，共 {len(files_info)} 个文件")
            return {"success": True, "files": files_info}

        except Exception as e:
            logger.error(f"获取日志文件列表失败: {str(e)}")
            ctx.log_with_user('error', f"获取日志文件列表失败: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/admin/logs/export')
    def export_log_file(file: str, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """导出指定的日志文件"""
        import os
        from fastapi.responses import StreamingResponse

        try:
            if not file:
                raise HTTPException(status_code=400, detail="缺少文件参数")

            safe_name = os.path.basename(file)
            log_dir = os.path.abspath("logs")
            target_path = os.path.abspath(os.path.join(log_dir, safe_name))

            # 防止目录遍历
            if not target_path.startswith(log_dir):
                ctx.log_with_user('warning', f"尝试访问非法日志文件: {file}", admin_user)
                raise HTTPException(status_code=400, detail="非法的日志文件路径")

            if not os.path.exists(target_path):
                ctx.log_with_user('warning', f"日志文件不存在: {file}", admin_user)
                raise HTTPException(status_code=404, detail="日志文件不存在")

            ctx.log_with_user('info', f"导出日志文件: {safe_name}", admin_user)
            def iter_file(path: str):
                file_handle = open(path, 'rb')
                try:
                    while True:
                        chunk = file_handle.read(8192)
                        if not chunk:
                            break
                        yield chunk
                finally:
                    file_handle.close()

            headers = {
                "Content-Disposition": f'attachment; filename="{safe_name}"'
            }
            return StreamingResponse(
                iter_file(target_path),
                media_type='text/plain; charset=utf-8',
                headers=headers
            )

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"导出日志文件失败: {str(e)}")
            ctx.log_with_user('error', f"导出日志文件失败: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/admin/stats')
    def get_system_stats(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """获取系统统计信息（管理员专用）"""
        from db_manager import db_manager
        try:
            ctx.log_with_user('info', "查询系统统计信息", admin_user)

            stats = {
                "users": {
                    "total": 0,
                    "active_today": 0
                },
                "cookies": {
                    "total": 0,
                    "enabled": 0
                },
                "cards": {
                    "total": 0,
                    "enabled": 0
                },
                "system": {
                    "uptime": "未知",
                    "version": "1.0.0"
                }
            }

            # 用户统计
            all_users = ctx.db_manager.get_all_users()
            stats["users"]["total"] = len(all_users)

            # Cookie统计
            all_cookies = ctx.db_manager.get_all_cookies()
            stats["cookies"]["total"] = len(all_cookies)

            # 卡券统计
            all_cards = ctx.db_manager.get_all_cards()
            if all_cards:
                stats["cards"]["total"] = len(all_cards)
                stats["cards"]["enabled"] = len([card for card in all_cards if card.get('enabled', True)])

            ctx.log_with_user('info', "系统统计信息查询完成", admin_user)
            return stats

        except Exception as e:
            ctx.log_with_user('error', f"获取系统统计信息失败: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/admin/backup/download')
    def download_database_backup(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """下载数据库备份文件（管理员专用）"""
        import os
        from fastapi.responses import FileResponse
        from datetime import datetime

        try:
            ctx.log_with_user('info', "请求下载数据库备份", admin_user)

            # 使用db_manager的实际数据库路径
            from db_manager import db_manager
            db_file_path = ctx.db_manager.db_path

            # 检查数据库文件是否存在
            if not os.path.exists(db_file_path):
                ctx.log_with_user('error', f"数据库文件不存在: {db_file_path}", admin_user)
                raise HTTPException(status_code=404, detail="数据库文件不存在")

            # 生成带时间戳的文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            download_filename = f"xianyu_backup_{timestamp}.db"

            ctx.log_with_user('info', f"开始下载数据库备份: {download_filename}", admin_user)

            return FileResponse(
                path=db_file_path,
                filename=download_filename,
                media_type='application/octet-stream'
            )

        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"下载数据库备份失败: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.post('/admin/backup/upload')
    async def upload_database_backup(admin_user: Dict[str, Any] = Depends(ctx.require_admin),
                                    backup_file: UploadFile = File(...)):
        """Validate and atomically restore a database backup."""
        import shutil
        from db_manager.base import validate_backup_database

        filename = backup_file.filename or ""
        if ctx.Path(filename).suffix.lower() != ".db":
            raise HTTPException(status_code=400, detail="只支持.db格式的数据库文件")

        current_db_path = ctx.Path(str(ctx.db_manager.db_path)).resolve()
        if str(ctx.db_manager.db_path) == ":memory:":
            raise HTTPException(status_code=400, detail="内存数据库不支持文件恢复")
        current_db_path.parent.mkdir(parents=True, exist_ok=True)
        temp_file_path = current_db_path.parent / f".restore-{ctx.uuid.uuid4().hex}.db"
        rollback_stage_path = current_db_path.parent / f".rollback-{ctx.uuid.uuid4().hex}.db"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_current_path = current_db_path.parent / f"xianyu_data_backup_{timestamp}.db"
        max_size = 100 * 1024 * 1024
        total_size = 0

        try:
            with temp_file_path.open("wb") as temp_file:
                while chunk := await backup_file.read(1024 * 1024):
                    total_size += len(chunk)
                    if total_size > max_size:
                        raise HTTPException(status_code=400, detail="备份文件大小不能超过100MB")
                    temp_file.write(chunk)

            try:
                validate_backup_database(temp_file_path)
            except ValueError:
                raise HTTPException(status_code=400, detail="无效或不完整的数据库文件")

            maintenance_lock = ctx.app.state.maintenance_lock
            if maintenance_lock.locked():
                raise HTTPException(status_code=503, detail="数据库维护正在进行")

            async with maintenance_lock:
                ctx.app.state.maintenance_mode = True
                runtime_manager = ctx.cookie_manager.manager
                paused = False
                try:
                    if runtime_manager is not None and hasattr(runtime_manager, "pause_for_maintenance"):
                        await runtime_manager.pause_for_maintenance()
                        paused = True

                    with ctx.db_manager.lock:
                        ctx.db_manager.close()
                        if not current_db_path.exists():
                            raise RuntimeError("live database does not exist")
                        shutil.copy2(current_db_path, backup_current_path)
                        os.replace(temp_file_path, current_db_path)
                        ctx.db_manager.reinitialize()

                    test_users = ctx.db_manager.get_all_users()
                    if runtime_manager is not None and hasattr(runtime_manager, "reload_from_db"):
                        runtime_manager.reload_from_db()

                    if paused and runtime_manager is not None and hasattr(runtime_manager, "resume_after_maintenance"):
                        await runtime_manager.resume_after_maintenance()
                        paused = False

                    ctx.SESSION_TOKENS.clear()
                    ctx.DOWNLOAD_TOKENS.clear()
                    ctx.log_with_user('info', f"数据库恢复成功，包含 {len(test_users)} 个用户", admin_user)
                    return {
                        "success": True,
                        "message": "数据库恢复成功，所有旧会话已失效",
                        "user_count": len(test_users),
                    }
                except Exception as restore_error:
                    logger.error(f"数据库恢复失败，开始回滚: {type(restore_error).__name__}")
                    try:
                        with ctx.db_manager.lock:
                            ctx.db_manager.close()
                            if backup_current_path.exists():
                                shutil.copy2(backup_current_path, rollback_stage_path)
                                os.replace(rollback_stage_path, current_db_path)
                            ctx.db_manager.reinitialize()
                        if runtime_manager is not None and hasattr(runtime_manager, "reload_from_db"):
                            runtime_manager.reload_from_db()
                    except Exception as rollback_error:
                        logger.critical(f"数据库回滚失败: {type(rollback_error).__name__}")
                        raise HTTPException(status_code=500, detail="数据库恢复与回滚均失败")
                    raise HTTPException(status_code=500, detail="数据库恢复失败，原数据库已恢复")
                finally:
                    if paused and runtime_manager is not None and hasattr(runtime_manager, "resume_after_maintenance"):
                        try:
                            await runtime_manager.resume_after_maintenance()
                        except Exception as resume_error:
                            logger.critical(f"账号任务恢复失败: {type(resume_error).__name__}")
                    ctx.app.state.maintenance_mode = False
        except HTTPException:
            raise
        except Exception as exc:
            logger.error(f"上传数据库备份失败: {type(exc).__name__}")
            raise HTTPException(status_code=500, detail="数据库恢复失败")
        finally:
            for cleanup_path in (temp_file_path, rollback_stage_path):
                try:
                    cleanup_path.unlink(missing_ok=True)
                except OSError:
                    pass

    @router.get('/admin/backup/list')
    def list_backup_files(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """列出服务器上的备份文件（管理员专用）"""
        import os
        import glob
        from datetime import datetime

        try:
            ctx.log_with_user('info', "查询备份文件列表", admin_user)

            # 查找备份文件（在data目录中）
            backup_files = glob.glob("data/xianyu_data_backup_*.db")

            backup_list = []
            for file_path in backup_files:
                try:
                    stat = os.stat(file_path)
                    backup_list.append({
                        'filename': os.path.basename(file_path),
                        'size': stat.st_size,
                        'size_mb': round(stat.st_size / (1024 * 1024), 2),
                        'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                        'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    })
                except Exception as e:
                    ctx.log_with_user('warning', f"读取备份文件信息失败: {file_path} - {str(e)}", admin_user)

            # 按修改时间倒序排列
            backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

            ctx.log_with_user('info', f"找到 {len(backup_list)} 个备份文件", admin_user)

            return {
                "backups": backup_list,
                "total": len(backup_list)
            }

        except Exception as e:
            ctx.log_with_user('error', f"查询备份文件列表失败: {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/admin/data/{table_name}')
    def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """获取指定表的所有数据（管理员专用）"""
        from db_manager import db_manager
        try:
            ctx.log_with_user('info', f"查询表数据: {table_name}", admin_user)

            # 验证表名安全性
            allowed_tables = [
                'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
                'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
                'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
                'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay"
            ]

            if table_name not in allowed_tables:
                ctx.log_with_user('warning', f"尝试访问不允许的表: {table_name}", admin_user)
                raise HTTPException(status_code=400, detail="不允许访问该表")

            # 获取表数据
            data, columns = ctx.db_manager.get_table_data(table_name)
            data = ctx._redact_admin_table_data(table_name, data, columns)

            ctx.log_with_user('info', f"表 {table_name} 查询成功，共 {len(data)} 条记录", admin_user)

            return {
                "success": True,
                "data": data,
                "columns": columns,
                "count": len(data)
            }

        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"查询表数据失败: {table_name} - {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/admin/data/{table_name}/export')
    def export_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """导出指定表的数据为Excel文件（管理员专用）"""
        from db_manager import db_manager
        import io
        try:
            ctx.log_with_user('info', f"导出表数据: {table_name}", admin_user)

            # 验证表名安全性
            allowed_tables = [
                'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
                'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
                'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
                'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
                'risk_control_logs'
            ]

            if table_name not in allowed_tables:
                ctx.log_with_user('warning', f"尝试导出不允许的表: {table_name}", admin_user)
                raise HTTPException(status_code=400, detail="不允许导出该表")

            # 获取表数据
            data, columns = ctx.db_manager.get_table_data(table_name)
            data = ctx._redact_admin_table_data(table_name, data, columns)

            if not data:
                raise HTTPException(status_code=400, detail="表中没有数据")

            # 创建Excel文件
            import openpyxl
            from openpyxl.utils import get_column_letter
        
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name

            # 写入表头
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, '')
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')

            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            ctx.log_with_user('info', f"表 {table_name} 导出成功，共 {len(data)} 条记录", admin_user)

            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={table_name}_export.xlsx"}
            )

        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"导出表数据失败: {table_name} - {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.delete('/admin/data/{table_name}/{record_id}')
    def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """删除指定表的指定记录（管理员专用）"""
        from db_manager import db_manager
        try:
            ctx.log_with_user('info', f"删除表记录: {table_name}.{record_id}", admin_user)

            # 验证表名安全性
            allowed_tables = [
                'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
                'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
                'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
                'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
            ]

            if table_name not in allowed_tables:
                ctx.log_with_user('warning', f"尝试删除不允许的表记录: {table_name}", admin_user)
                raise HTTPException(status_code=400, detail="不允许操作该表")

            # 特殊保护：不能删除管理员用户
            if table_name == 'users' and record_id == str(admin_user['user_id']):
                ctx.log_with_user('warning', "尝试删除管理员自己", admin_user)
                raise HTTPException(status_code=400, detail="不能删除管理员自己")

            # 删除记录
            success = ctx.db_manager.delete_table_record(table_name, record_id)

            if success:
                ctx.log_with_user('info', f"表记录删除成功: {table_name}.{record_id}", admin_user)
                return {"success": True, "message": "删除成功"}
            else:
                ctx.log_with_user('warning', f"表记录删除失败: {table_name}.{record_id}", admin_user)
                raise HTTPException(status_code=400, detail="删除失败，记录可能不存在")

        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"删除表记录异常: {table_name}.{record_id} - {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.delete('/admin/data/{table_name}')
    def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """清空指定表的所有数据（管理员专用）"""
        from db_manager import db_manager
        try:
            ctx.log_with_user('info', f"清空表数据: {table_name}", admin_user)

            # 验证表名安全性
            allowed_tables = [
                'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
                'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
                'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
                'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
                'risk_control_logs'
            ]

            # 不允许清空用户表
            if table_name == 'users':
                ctx.log_with_user('warning', "尝试清空用户表", admin_user)
                raise HTTPException(status_code=400, detail="不允许清空用户表")

            if table_name not in allowed_tables:
                ctx.log_with_user('warning', f"尝试清空不允许的表: {table_name}", admin_user)
                raise HTTPException(status_code=400, detail="不允许清空该表")

            # 清空表数据
            success = ctx.db_manager.clear_table_data(table_name)

            if success:
                ctx.log_with_user('info', f"表数据清空成功: {table_name}", admin_user)
                return {"success": True, "message": "清空成功"}
            else:
                ctx.log_with_user('warning', f"表数据清空失败: {table_name}", admin_user)
                raise HTTPException(status_code=400, detail="清空失败")

        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"清空表数据异常: {table_name} - {str(e)}", admin_user)
            raise HTTPException(status_code=500, detail=str(e))

    @router.get("/backup/export")
    def export_backup(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """导出用户备份"""
        try:
            from db_manager import db_manager
            user_id = current_user['user_id']
            username = current_user['username']

            # 导出当前用户的数据
            backup_data = ctx.db_manager.export_backup(user_id)

            # 生成文件名
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"xianyu_backup_{username}_{timestamp}.json"

            # 返回JSON响应，设置下载头
            response = JSONResponse(content=backup_data)
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-Type"] = "application/json"

            return response
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"导出备份失败: {str(e)}")

    @router.post("/backup/import")
    def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """导入用户备份"""
        try:
            # 验证文件类型
            if not file.filename.endswith('.json'):
                raise HTTPException(status_code=400, detail="只支持JSON格式的备份文件")

            # 读取文件内容
            content = file.file.read()
            backup_data = json.loads(content.decode('utf-8'))

            # 导入备份到当前用户
            from db_manager import db_manager
            user_id = current_user['user_id']
            success = ctx.db_manager.import_backup(backup_data, user_id)

            if success:
                # 备份导入成功后，刷新 CookieManager 的内存缓存
                import cookie_manager
                if ctx.cookie_manager.manager:
                    try:
                        ctx.cookie_manager.manager.reload_from_db()
                        logger.info("备份导入后已刷新 CookieManager 缓存")
                    except Exception as e:
                        logger.error(f"刷新 CookieManager 缓存失败: {e}")

                return {"message": "备份导入成功"}
            else:
                raise HTTPException(status_code=400, detail="备份导入失败")

        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="备份文件格式无效")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"导入备份失败: {str(e)}")

    @router.post("/system/reload-cache")
    def reload_cache(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        """重新加载系统缓存（用于手动刷新数据）"""
        try:
            import cookie_manager
            if ctx.cookie_manager.manager:
                success = ctx.cookie_manager.manager.reload_from_db()
                if success:
                    return {"message": "系统缓存已刷新", "success": True}
                else:
                    raise HTTPException(status_code=500, detail="缓存刷新失败")
            else:
                raise HTTPException(status_code=500, detail="CookieManager 未初始化")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"刷新缓存失败: {str(e)}")

    @router.get('/api/update/check')
    async def check_for_updates(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        检查是否有可用更新
    
        返回更新信息，包括新版本号、更新内容等
        """
        try:
            updater = ctx.get_updater()
            manifest = await updater.check_for_updates()
        
            if manifest is None:
                return {
                    "success": True,
                    "data": {
                        "has_update": False,
                        "current_version": updater.current_version,
                        "message": "已是最新版本"
                    }
                }
        
            # 获取需要更新的文件
            files_to_update = await updater.get_files_to_update(manifest)
            files_to_delete = await updater.get_files_to_delete(manifest)
            total_size = sum(f.size for f in files_to_update)

            if not files_to_update and not files_to_delete:
                return {
                    "success": True,
                    "data": {
                        "has_update": False,
                        "current_version": updater.current_version,
                        "message": "已是最新版本"
                    }
                }
        
            return {
                "success": True,
                "data": {
                    "has_update": True,
                    "current_version": updater.current_version,
                    "new_version": manifest.version,
                    "description": manifest.description,
                    "changelog": manifest.changelog or [],
                    "files_count": len(files_to_update),
                    "deleted_files_count": len(files_to_delete),
                    "total_size": total_size,
                    "release_date": manifest.release_date,
                    "files": [
                        {
                            "path": f.path,
                            "size": f.size,
                            "requires_restart": f.requires_restart,
                            "description": f.description
                        }
                        for f in files_to_update
                    ],
                    "deleted_files": [
                        {
                            "path": f.path,
                            "requires_restart": f.requires_restart,
                            "description": f.description
                        }
                        for f in files_to_delete
                    ]
                }
            }
        
        except Exception as e:
            logger.error(f"检查更新失败: {e}")
            return {
                "success": False,
                "message": f"检查更新失败: {str(e)}"
            }

    @router.post('/api/update/apply')
    async def apply_updates(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        应用更新
    
        下载并安装所有可用更新
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            updater = ctx.get_updater()
        
            ctx.log_with_user('info', "开始执行自动更新", current_user)
        
            result = await updater.perform_update()
        
            if result["success"]:
                ctx.log_with_user('info', f"更新完成: {result['message']}", current_user)
            else:
                ctx.log_with_user('error', f"更新失败: {result['message']}", current_user)
        
            return {
                "success": result["success"],
                "data": result
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"应用更新失败: {e}")
            return {
                "success": False,
                "message": f"应用更新失败: {str(e)}"
            }

    @router.get('/api/update/progress')
    async def get_update_progress(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        获取更新进度
    
        返回当前更新状态和进度信息
        """
        try:
            updater = ctx.get_updater()
            progress = updater.progress
        
            return {
                "success": True,
                "data": {
                    "status": progress.status.value,
                    "current_file": progress.current_file,
                    "current_index": progress.current_index,
                    "total_files": progress.total_files,
                    "downloaded_bytes": progress.downloaded_bytes,
                    "total_bytes": progress.total_bytes,
                    "message": progress.message,
                    "error": progress.error
                }
            }
        
        except Exception as e:
            logger.error(f"获取更新进度失败: {e}")
            return {
                "success": False,
                "message": f"获取更新进度失败: {str(e)}"
            }

    @router.get('/api/update/local-hashes')
    async def get_local_file_hashes(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        获取本地文件哈希值
    
        用于服务端比对哪些文件需要更新
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            updater = ctx.get_updater()
            hashes = updater.get_local_file_hashes()
        
            return {
                "success": True,
                "data": {
                    "version": updater.current_version,
                    "files": hashes,
                    "count": len(hashes)
                }
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"获取文件哈希失败: {e}")
            return {
                "success": False,
                "message": f"获取文件哈希失败: {str(e)}"
            }

    @router.post('/api/update/cleanup-backups')
    async def cleanup_old_backups(days: int = 7, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        清理旧的备份文件
    
        Args:
            days: 保留天数，默认7天
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            updater = ctx.get_updater()
            updater.cleanup_old_backups(keep_days=days)
        
            ctx.log_with_user('info', f"清理了 {days} 天前的备份文件", current_user)
        
            return {
                "success": True,
                "message": f"已清理 {days} 天前的备份文件"
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"清理备份失败: {e}")
            return {
                "success": False,
                "message": f"清理备份失败: {str(e)}"
            }

    @router.get('/api/update/file-changes')
    async def get_file_changes(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        比较当前文件与上次更新后的哈希清单
    
        用于检测哪些文件在更新后被本地修改过
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            updater = ctx.get_updater()
            result = updater.compare_file_hashes()
        
            return {
                "success": True,
                "data": result
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"比较文件变化失败: {e}")
            return {
                "success": False,
                "message": f"比较文件变化失败: {str(e)}"
            }

    @router.post('/api/update/save-hashes')
    async def save_current_hashes(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        手动保存当前文件的哈希清单
    
        用于记录当前状态，以便以后比较
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            updater = ctx.get_updater()
            updater.save_file_hashes(updater.current_version)
        
            ctx.log_with_user('info', "手动保存文件哈希清单", current_user)
        
            return {
                "success": True,
                "message": "文件哈希清单已保存"
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"保存哈希清单失败: {e}")
            return {
                "success": False,
                "message": f"保存哈希清单失败: {str(e)}"
            }

    @router.get('/api/update/saved-hashes')
    async def get_saved_hashes(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        获取上次保存的文件哈希清单
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            updater = ctx.get_updater()
            saved_hashes = updater.load_file_hashes()
        
            if saved_hashes is None:
                return {
                    "success": True,
                    "data": None,
                    "message": "没有保存的哈希清单"
                }
        
            return {
                "success": True,
                "data": {
                    "version": saved_hashes.get("version"),
                    "updated_at": saved_hashes.get("updated_at"),
                    "total_files": saved_hashes.get("total_files"),
                    "last_updated_files": saved_hashes.get("last_updated_files", []),
                    "last_updated_count": saved_hashes.get("last_updated_count", 0)
                }
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"获取哈希清单失败: {e}")
            return {
                "success": False,
                "message": f"获取哈希清单失败: {str(e)}"
            }

    @router.post('/api/update/restart')
    async def restart_application(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """
        重启应用（用于更新后重启）
    
        注意：此操作会重启整个应用
        """
        try:
            ctx._ensure_update_admin(current_user)
        
            ctx.log_with_user('info', "用户请求重启应用", current_user)
        
            import subprocess
            import sys
        
            # 返回响应后异步重启
            async def delayed_restart():
                await asyncio.sleep(2)  # 等待2秒让响应返回
                logger.info("正在重启应用...")
            
                # 获取当前Python解释器和脚本路径
                python = sys.executable
                script = sys.argv[0]
            
                # 在Windows上使用start命令启动新进程
                if sys.platform == 'win32':
                    subprocess.Popen(
                        [python, script],
                        creationflags=subprocess.CREATE_NEW_CONSOLE
                    )
                else:
                    # Linux/Mac
                    subprocess.Popen([python, script])
            
                # 退出当前进程
                os._exit(0)
        
            # 创建后台任务
            asyncio.create_task(delayed_restart())
        
            return {
                "success": True,
                "message": "应用将在2秒后重启"
            }
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"重启应用失败: {e}")
            return {
                "success": False,
                "message": f"重启应用失败: {str(e)}"
            }

    @router.post("/accounts/{cid}/polish-items")
    async def polish_account_items(cid: str, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """擦亮指定账号的所有在售商品"""
        try:
            cid = ctx._ensure_cookie_access(cid, current_user)
            cookie_info = ctx.db_manager.get_cookie_by_id(cid)
            if not cookie_info:
                return {"success": False, "message": "未找到指定的账号信息"}

            cookies_str = cookie_info.get('cookies_str', '')
            if not cookies_str:
                return {"success": False, "message": "账号cookie信息为空"}

            from XianyuAutoAsync import XianyuLive
            xianyu_instance = XianyuLive(cookies_str, cid, register_instance=False)

            logger.info(f"开始擦亮账号 {cid} 的所有商品")
            result = await xianyu_instance.polish_all_items()

            await xianyu_instance.close_session()

            return result

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"擦亮账号商品异常: {str(e)}")
            return {"success": False, "message": f"擦亮异常: {str(e)}"}

    @router.post("/scheduled-tasks")
    async def create_scheduled_task(request: dict, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """创建定时任务"""
        try:
            account_id = request.get('account_id', '').strip()
            run_hour = ctx._parse_run_hour(request.get('run_hour', request.get('delay_minutes', 8)))
            random_delay_max = ctx._parse_random_delay(request.get('random_delay_max', 10), 10)
            enabled = ctx._parse_enabled_flag(request.get('enabled', True))

            if not account_id:
                return {"success": False, "message": "账号ID不能为空"}

            cookie_details = ctx.db_manager.get_cookie_details(account_id)
            if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
                return {"success": False, "message": "账号不存在或无权创建此任务"}

            name = f"每日擦亮-{account_id}"
            next_run_at = ctx.db_manager.calculate_next_daily_run(run_hour, random_delay_max, include_today=True)

            existing_task = ctx.db_manager.get_scheduled_task_by_account(
                account_id,
                user_id=current_user['user_id'],
                task_type='item_polish'
            )

            if existing_task:
                updated = ctx.db_manager.update_scheduled_task(
                    existing_task['id'],
                    name=name,
                    interval_hours=24,
                    delay_minutes=run_hour,
                    random_delay_max=random_delay_max,
                    enabled=enabled,
                    next_run_at=next_run_at
                )
                if updated:
                    task = ctx.db_manager.get_scheduled_task(existing_task['id'])
                    return {
                        "success": True,
                        "message": "定时擦亮任务更新成功",
                        "task_id": existing_task['id'],
                        "task": task
                    }
                return {"success": False, "message": "更新定时任务失败"}

            task_id = ctx.db_manager.create_scheduled_task(
                name=name, task_type='item_polish', account_id=account_id,
                user_id=current_user['user_id'],
                interval_hours=24, delay_minutes=run_hour,
                random_delay_max=random_delay_max,
                next_run_at=next_run_at,
                enabled=enabled
            )

            if task_id:
                task = ctx.db_manager.get_scheduled_task(task_id)
                return {"success": True, "message": "定时擦亮任务创建成功", "task_id": task_id, "task": task}
            else:
                return {"success": False, "message": "创建定时任务失败"}
        except Exception as e:
            logger.error(f"创建定时任务异常: {str(e)}")
            return {"success": False, "message": f"创建定时任务异常: {str(e)}"}

    @router.get("/scheduled-tasks")
    async def list_scheduled_tasks(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """获取定时任务列表"""
        try:
            tasks = ctx.db_manager.get_scheduled_tasks(user_id=current_user['user_id'])
            return {"success": True, "tasks": tasks}
        except Exception as e:
            logger.error(f"获取定时任务列表异常: {str(e)}")
            return {"success": False, "message": f"获取定时任务列表异常: {str(e)}"}

    @router.put("/scheduled-tasks/{task_id}")
    async def update_scheduled_task(task_id: int, request: dict, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """更新定时任务"""
        try:
            task = ctx.db_manager.get_scheduled_task(task_id)
            if not task:
                return {"success": False, "message": "任务不存在"}
            if task['user_id'] != current_user['user_id']:
                return {"success": False, "message": "无权修改此任务"}

            kwargs = {}

            if 'name' in request:
                name = str(request.get('name') or '').strip()
                if name:
                    kwargs['name'] = name

            if 'interval_hours' in request:
                kwargs['interval_hours'] = int(request.get('interval_hours', task.get('interval_hours', 24)))

            if 'run_hour' in request or 'delay_minutes' in request:
                kwargs['delay_minutes'] = ctx._parse_run_hour(request.get('run_hour', request.get('delay_minutes')))

            if 'random_delay_max' in request:
                kwargs['random_delay_max'] = ctx._parse_random_delay(
                    request.get('random_delay_max'),
                    task.get('random_delay_max', 10)
                )

            if 'enabled' in request:
                kwargs['enabled'] = ctx._parse_enabled_flag(request.get('enabled'))

            effective_enabled = kwargs.get('enabled', 1 if task['enabled'] else 0)
            effective_run_hour = kwargs.get('delay_minutes', task.get('delay_minutes', 8))
            effective_random_delay = kwargs.get('random_delay_max', task.get('random_delay_max', 10))

            if task['task_type'] == 'item_polish' and effective_enabled:
                should_reschedule = (
                    'delay_minutes' in kwargs or
                    'random_delay_max' in kwargs or
                    ('enabled' in kwargs and not task['enabled'])
                )
                if should_reschedule:
                    kwargs['next_run_at'] = ctx.db_manager.calculate_next_daily_run(
                        effective_run_hour,
                        effective_random_delay,
                        include_today=True
                    )

            if not kwargs:
                return {"success": False, "message": "没有可更新的字段"}

            if ctx.db_manager.update_scheduled_task(task_id, **kwargs):
                updated_task = ctx.db_manager.get_scheduled_task(task_id)
                return {"success": True, "message": "定时任务更新成功", "task": updated_task}
            else:
                return {"success": False, "message": "更新失败"}
        except Exception as e:
            logger.error(f"更新定时任务异常: {str(e)}")
            return {"success": False, "message": f"更新定时任务异常: {str(e)}"}

    @router.delete("/scheduled-tasks/{task_id}")
    async def delete_scheduled_task(task_id: int, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """删除定时任务"""
        try:
            task = ctx.db_manager.get_scheduled_task(task_id)
            if not task:
                return {"success": False, "message": "任务不存在"}
            if task['user_id'] != current_user['user_id']:
                return {"success": False, "message": "无权删除此任务"}

            if ctx.db_manager.delete_scheduled_task(task_id):
                return {"success": True, "message": "定时任务已删除"}
            else:
                return {"success": False, "message": "删除失败"}
        except Exception as e:
            logger.error(f"删除定时任务异常: {str(e)}")
            return {"success": False, "message": f"删除定时任务异常: {str(e)}"}

    @router.put("/scheduled-tasks/{task_id}/toggle")
    async def toggle_scheduled_task(task_id: int, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """启用/禁用定时任务"""
        try:
            task = ctx.db_manager.get_scheduled_task(task_id)
            if not task:
                return {"success": False, "message": "任务不存在"}
            if task['user_id'] != current_user['user_id']:
                return {"success": False, "message": "无权操作此任务"}

            new_enabled = 0 if task['enabled'] else 1
            update_kwargs = {'enabled': new_enabled}
            if new_enabled:
                update_kwargs['next_run_at'] = ctx.db_manager.calculate_next_daily_run(
                    task.get('delay_minutes', 8),
                    task.get('random_delay_max', 10),
                    include_today=True
                )

            if ctx.db_manager.update_scheduled_task(task_id, **update_kwargs):
                status = "启用" if new_enabled else "禁用"
                updated_task = ctx.db_manager.get_scheduled_task(task_id)
                return {
                    "success": True,
                    "message": f"定时任务已{status}",
                    "enabled": bool(new_enabled),
                    "task": updated_task
                }
            else:
                return {"success": False, "message": "操作失败"}
        except Exception as e:
            logger.error(f"切换定时任务状态异常: {str(e)}")
            return {"success": False, "message": f"操作异常: {str(e)}"}

    @router.post("/api/analytics/error")
    async def report_client_error(req: ctx.ClientErrorRequest):
        """???? JS ????"""
        ctx.track(
            user="browser",
            action="client_error",
            target=req.source.split("/")[-1] if req.source else "-",
            result="error",
            detail="L{}:C{} {} | {}".format(req.lineno, req.colno, req.message[:100], req.url)
        )
        return {"success": True}

    @router.get("/api/files")
    async def list_files(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        try:
            user_id = current_user['user_id']
            files = ctx.db_manager.get_files(user_id=user_id)
            ctx.track(user=current_user.get("username","?"), action=ctx.ActionEvent.FILE_LIST, detail="count={}".format(len(files)))
            return {"success": True, "data": files}
        except Exception as e:
            logger.error("list_files failed: {}".format(e))
            return {"success": False, "message": str(e)}

    @router.get("/api/files/{file_id}/download")
    async def download_file(file_id: int, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        try:
            user_id = current_user['user_id']
            can_download, remaining, max_allowed = ctx.db_manager.check_download_quota(file_id, user_id)
            if not can_download:
                ctx.track(user=current_user.get("username","?"), action=ctx.ActionEvent.FILE_DOWNLOAD, target=str(file_id), result="quota_exceeded", detail="max={}".format(max_allowed))
                raise HTTPException(status_code=403, detail="下载次数已用完")
            file_info = ctx.db_manager.get_file(file_id)
            if not file_info:
                raise HTTPException(status_code=404, detail="文件不存在")
            ctx.db_manager.record_download(file_id, user_id)
            ctx.track(user=str(user_id), action=ctx.ActionEvent.FILE_DOWNLOAD, target="{}:{}".format(file_id, file_info.get("filename","?")), detail="via=token")
            return Response(
                content=file_info['file_data'],
                media_type=file_info.get('mime_type', 'application/octet-stream'),
                headers={"Content-Disposition": "attachment; filename*=UTF-8''" + urllib.parse.quote(file_info['filename'])}
            )
        except HTTPException:
            raise
        except Exception as e:
            logger.error("download_file failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.post("/api/files")
    async def upload_file(
        file: UploadFile = File(...),
        description: str = Form(""),
        max_downloads: int = Form(5),
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        try:
            file_data = await file.read()
            file_id = ctx.db_manager.add_file(
                filename=file.filename,
                file_data=file_data,
                description=description,
                mime_type=file.content_type or 'application/octet-stream',
                max_downloads=max_downloads,
                created_by=admin_user['user_id']
            )
            logger.info("admin {} uploaded file: {} (id={})".format(admin_user['username'], file.filename, file_id))
            ctx.track(user=admin_user.get("username","?"), action=ctx.ActionEvent.FILE_UPLOAD, target="{}:{}".format(file_id, file.filename), detail="size={}".format(len(file_data)))
            return {"success": True, "message": "文件上传成功", "file_id": file_id}
        except Exception as e:
            logger.error("upload_file failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.put("/api/files/{file_id}")
    async def update_file_info(
        file_id: int,
        description: Optional[str] = Form(None),
        max_downloads: Optional[int] = Form(None),
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        try:
            success = ctx.db_manager.update_file(file_id, description=description, max_downloads=max_downloads)
            if not success:
                raise HTTPException(status_code=404, detail="文件不存在或更新失败")
            ctx.track(user=admin_user.get("username","?"), action=ctx.ActionEvent.FILE_EDIT, target=str(file_id), detail="desc={},max_dl={}".format(description, max_downloads))
            return {"success": True, "message": "文件更新成功"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error("update_file_info failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.delete("/api/files/{file_id}")
    async def delete_file_info(
        file_id: int,
        admin_user: Dict[str, Any] = Depends(ctx.require_admin)
    ):
        try:
            success = ctx.db_manager.delete_file(file_id)
            if not success:
                raise HTTPException(status_code=404, detail="文件不存在")
            return {"success": True, "message": "文件已删除"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error("delete_file_info failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.get("/api/files/{file_id}/download-token")
    async def get_download_token(file_id: int, current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        try:
            user_id = current_user['user_id']
            can_download, remaining, max_allowed = ctx.db_manager.check_download_quota(file_id, user_id)
            if not can_download:
                ctx.track(user=current_user.get("username","?"), action=ctx.ActionEvent.FILE_DOWNLOAD, target=str(file_id), result="quota_exceeded", detail="max={}".format(max_allowed))
                raise HTTPException(status_code=403, detail="下载次数已用完")
            token_str = secrets.token_urlsafe(32)
            ctx.DOWNLOAD_TOKENS[token_str] = {
                "user_id": user_id,
                "file_id": file_id,
                "exp": time.time() + 120
            }
            return {"success": True, "token": token_str}
        except HTTPException:
            raise
        except Exception as e:
            logger.error("get_download_token failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.get("/api/files/{file_id}/direct")
    async def direct_download(file_id: int, token: str = None):
        try:
            if not token or token not in ctx.DOWNLOAD_TOKENS:
                raise HTTPException(status_code=401, detail="无效或已过期的下载链接")
            token_data = ctx.DOWNLOAD_TOKENS[token]
            if token_data["file_id"] != file_id or time.time() > token_data["exp"]:
                del ctx.DOWNLOAD_TOKENS[token]
                raise HTTPException(status_code=401, detail="下载链接已过期")
            user_id = token_data["user_id"]
            del ctx.DOWNLOAD_TOKENS[token]  # 一次性
            file_info = ctx.db_manager.get_file(file_id)
            if not file_info:
                raise HTTPException(status_code=404, detail="文件不存在")
            ctx.db_manager.record_download(file_id, user_id)
            ctx.track(user=str(user_id), action=ctx.ActionEvent.FILE_DOWNLOAD, target="{}:{}".format(file_id, file_info.get("filename","?")), detail="via=token")
            return Response(
                content=file_info['file_data'],
                media_type=file_info.get('mime_type', 'application/octet-stream'),
                headers={"Content-Disposition": "attachment; filename*=UTF-8''" + urllib.parse.quote(file_info['filename'])}
            )
        except HTTPException:
            raise
        except Exception as e:
            logger.error("direct_download failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.post("/api/groups")
    async def create_group(req: ctx.CreateGroupRequest, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        try:
            group_id = ctx.db_manager.create_group(
                group_name=req.group_name,
                description=req.description or "",
                created_by=admin_user["user_id"]
            )
            members = ctx.db_manager.batch_create_users(group_id, req.user_count, req.group_name)
            logger.info("admin {} created group '{}' with {} users".format(admin_user["username"], req.group_name, len(members)))
            ctx.track(user=admin_user.get("username","?"), action=ctx.ActionEvent.GROUP_CREATE, target="{}:{}".format(group_id, req.group_name), detail="members={}".format(len(members)))
            return {
                "success": True,
                "message": "用户组 '{}' 创建成功，含 {} 个用户".format(req.group_name, len(members)),
                "group": {"id": group_id, "group_name": req.group_name},
                "members": members
            }
        except Exception as e:
            logger.error("create_group failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.get("/api/groups")
    async def list_groups(admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        try:
            groups = ctx.db_manager.get_all_groups()
            return {"success": True, "data": groups}
        except Exception as e:
            logger.error("list_groups failed: {}".format(e))
            return {"success": False, "message": str(e)}

    @router.get("/api/groups/{group_id}/members")
    async def get_group_members(group_id: int, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        try:
            members = ctx.db_manager.get_group_members(group_id)
            return {"success": True, "data": members}
        except Exception as e:
            logger.error("get_group_members failed: {}".format(e))
            return {"success": False, "message": str(e)}

    @router.delete("/api/groups/{group_id}")
    async def delete_group(group_id: int, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        try:
            success = ctx.db_manager.delete_group(group_id)
            if not success:
                raise HTTPException(status_code=404, detail="用户组不存在")
            logger.info("admin {} deleted group {}".format(admin_user["username"], group_id))
            return {"success": True, "message": "用户组已删除"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error("delete_group failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.post("/api/groups/{group_id}/members")
    async def add_group_members(group_id: int, req: ctx.AddMembersRequest, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        try:
            members = ctx.db_manager.batch_create_users(group_id, req.count, "grp{}".format(group_id))
            return {"success": True, "message": "已添加 {} 个用户".format(len(members)), "members": members}
        except Exception as e:
            logger.error("add_group_members failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.delete("/api/groups/{group_id}/members/{user_id}")
    async def remove_group_member(group_id: int, user_id: int, admin_user: Dict[str, Any] = Depends(ctx.require_admin)):
        try:
            success = ctx.db_manager.remove_group_member(user_id)
            if not success:
                raise HTTPException(status_code=404, detail="用户不存在或不属于任何组")
            return {"success": True, "message": "成员已删除"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error("remove_group_member failed: {}".format(e))
            raise HTTPException(status_code=500, detail=str(e))

    @router.get('/api/blacklist/personal')
    def get_personal_blacklist(
        buyer_id: str = None,
        buyer_nick: str = None,
        page: int = 1,
        page_size: int = 20,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            result = ctx.blacklist_service.list_personal(
                user_id=current_user['user_id'],
                buyer_id=buyer_id,
                buyer_nick=buyer_nick,
                page=page,
                page_size=page_size,
            )
            return {'success': True, **result}
        except Exception as e:
            ctx.log_with_user('error', f"查询个人黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='查询个人黑名单失败')

    @router.post('/api/blacklist/personal')
    def create_personal_blacklist(
        request: ctx.PersonalBlacklistCreateRequest,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            cookie_id = str(request.cookie_id or '').strip() or None
            if cookie_id:
                cookie_id = ctx._ensure_cookie_access(cookie_id, current_user)

            result = ctx.blacklist_service.create_personal(
                user_id=current_user['user_id'],
                buyer_ids=request.buyer_ids,
                cookie_id=cookie_id,
                item_id=str(request.item_id or '').strip() or None,
                reason=str(request.reason or '').strip(),
                is_enabled=bool(request.is_enabled),
                buyer_nick=str(request.buyer_nick or '').strip(),
            )
            created = int(result.get('created') or 0)
            skipped = int(result.get('skipped') or 0)
            message = f"成功添加 {created} 条黑名单"
            if skipped:
                message += f"，跳过 {skipped} 条"
            ctx.log_with_user('info', f"新增个人黑名单: created={created}, skipped={skipped}", current_user)
            return {
                'success': True,
                'message': message,
                'data': {
                    'count': created,
                    'skipped': skipped,
                    'records': result.get('records') or [],
                },
            }
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"新增个人黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='新增个人黑名单失败')

    @router.post('/api/blacklist/personal/batch-delete')
    def batch_delete_personal_blacklist(
        request: ctx.PersonalBlacklistBatchDeleteRequest,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            deleted = ctx.blacklist_service.batch_delete_personal(request.ids, current_user['user_id'])
            return {'success': True, 'message': f'成功删除 {deleted} 条黑名单', 'data': {'deleted': deleted}}
        except Exception as e:
            ctx.log_with_user('error', f"批量删除个人黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='批量删除个人黑名单失败')

    @router.delete('/api/blacklist/personal/{record_id}')
    def delete_personal_blacklist(
        record_id: int,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            success = ctx.blacklist_service.delete_personal(record_id, current_user['user_id'])
            if not success:
                raise HTTPException(status_code=404, detail='黑名单记录不存在')
            return {'success': True, 'message': '删除成功'}
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"删除个人黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='删除个人黑名单失败')

    @router.get('/api/blacklist/personal/export')
    def export_personal_blacklist(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        try:
            content = ctx.blacklist_service.export_personal_xlsx(current_user['user_id'])
            filename = f"personal_blacklist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
            return Response(
                content=content,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers=headers,
            )
        except Exception as e:
            ctx.log_with_user('error', f"导出个人黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='导出个人黑名单失败')

    @router.post('/api/blacklist/personal/import')
    async def import_personal_blacklist(
        file: UploadFile = File(...),
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            filename = file.filename or ''
            if not filename.lower().endswith('.xlsx'):
                raise HTTPException(status_code=400, detail='仅支持 .xlsx 文件')
            content = await file.read()
            result = ctx.blacklist_service.import_personal_xlsx(current_user['user_id'], content)
            return {
                'success': True,
                'message': f"导入完成：新增 {result.get('created', 0)} 条，跳过 {result.get('skipped', 0)} 条",
                'data': result,
            }
        except HTTPException:
            raise
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            ctx.log_with_user('error', f"导入个人黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='导入个人黑名单失败')

    @router.get('/api/blacklist/platform')
    def get_platform_blacklist(
        page: int = 1,
        page_size: int = 20,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            result = ctx.blacklist_service.list_platform(current_user['user_id'], page=page, page_size=page_size)
            return {'success': True, **result}
        except Exception as e:
            ctx.log_with_user('error', f"查询平台黑名单失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='查询平台黑名单失败')

    @router.get("/api/announcement")
    def get_dashboard_announcement(current_user: Dict[str, Any] = Depends(ctx.get_current_user)):
        """获取仪表盘公告，优先读取 GitHub 公告文件，本地文件兜底。"""
        try:
            _ = current_user['user_id']
            snapshot = ctx._get_dashboard_announcement_payload()
            return {
                'success': True,
                'announcement': snapshot.get('current'),
                'current': snapshot.get('current'),
                'history': snapshot.get('history') or [],
            }
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"获取仪表盘公告失败: {ctx.mask_sensitive_text(e)}")
            return {
                'success': False,
                'announcement': None,
                'current': None,
                'history': [],
                'message': ctx.safe_client_error("获取公告失败，请稍后重试"),
            }
    @router.patch('/api/blacklist/personal/{record_id}/toggle')
    def toggle_personal_blacklist(
        record_id: int,
        request: ctx.PersonalBlacklistToggleRequest,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            success = ctx.blacklist_service.toggle_personal(record_id, current_user['user_id'], request.is_enabled)
            if not success:
                raise HTTPException(status_code=404, detail='黑名单记录不存在')
            return {'success': True, 'message': '状态已更新'}
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"更新个人黑名单状态失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='更新个人黑名单状态失败')
    @router.patch('/api/blacklist/personal/{record_id}/toggle')
    def toggle_personal_blacklist(
        record_id: int,
        request: ctx.PersonalBlacklistToggleRequest,
        current_user: Dict[str, Any] = Depends(ctx.get_current_user),
    ):
        try:
            success = ctx.blacklist_service.toggle_personal(record_id, current_user['user_id'], request.is_enabled)
            if not success:
                raise HTTPException(status_code=404, detail='黑名单记录不存在')
            return {'success': True, 'message': '状态已更新'}
        except HTTPException:
            raise
        except Exception as e:
            ctx.log_with_user('error', f"更新个人黑名单状态失败: {ctx.mask_sensitive_text(e)}", current_user)
            raise HTTPException(status_code=500, detail='更新个人黑名单状态失败')
    return router
